import time
import datetime
import requests
import re
from bs4 import BeautifulSoup
from lxml.etree import tostring
import lxml.html
import openpyxl as ol


# 株価整形関数
def Str2Float(num):
    n = num.replace(",", "")
    return float(n)


# [ページ]　注目ニュース一覧
url = "https://kabutan.jp/news/marketnews/?category=9"
html = requests.get(url)
soup = BeautifulSoup(html.content, "html.parser")

# aタグのテキストを抽出
for element in soup.find_all("a"):
    # 「今朝」が含まれている記事のhref属性値をlinkに代入
    if "今朝" in element.text:
        link = element.get("href")
        break


# [ページ]　当日の注目ニュース
url = "https://kabutan.jp" + link
req = requests.get(url)
html = req.text
dom = lxml.html.fromstring(html)

# div class='body'のxpathを取得
target_xpath = '//*[@id="shijyounews"]/article/div'
scraped_data = dom.xpath(target_xpath)

# 銘柄名、銘柄URL、材料を取得
meigaralists, codelists, descriptions = [], [], []
for news in scraped_data:
    articles = tostring(news, encoding="utf-8").decode().replace("\n", "")
    # 空白行を削除
    articles = list(filter(lambda x: x != "", articles.split("<br/>")))
    # 【好材料】等の行を削除
    articles = list(filter(lambda x: "【" not in x, articles))
    # 4行目から2step毎の各情報を取得する
    for i, codelist in enumerate(articles[3:]):
        # サーバー負荷軽減
        time.sleep(1)
        # 1行ずつ表示確認
        print(i, codelist)
        # ループ終了判定
        if "※★は上昇" in codelist:
            break
        # 偶数列
        elif i % 2 == 0:
            # 銘柄名リストに追加
            meigaralists.append(codelist.split()[0])
            # コードURLリストに追加
            codelists.append("https://kabutan.jp" + codelist.split('"')[1])
        # 奇数列
        elif i % 2 == 1:
            # 材料リストに追加(コード部分を整形)
            descriptions.append(
                re.sub(
                    r'&lt;<a[^>]*">(.*?)</a>&gt;',
                    r"<\1>",
                    codelist,
                )
            )

# excelを読み込む処理
wb = ol.load_workbook("kabutan.xlsx")
ws = wb["Sheet1"]
print("write start----------------------------")

# [ページ]　銘柄
for i in range(len(codelists)):
    # サーバー負荷軽減
    time.sleep(1)
    url = codelists[i]
    req = requests.get(url)
    soup = BeautifulSoup(req.text, "html.parser")
    # 銘柄名抽出
    meigara = soup.select("#kobetsu_left dd")
    # 株価抽出
    price = soup.select("#kobetsu_left tr")
    try:
        # 前日終値の日付部分を削除し成形　他株価は成形のみ
        yesterday_endprice = Str2Float(meigara[0].getText()[:-8])
        openprice = Str2Float(price[0].select("td")[0].getText())
        highprice = Str2Float(price[1].select("td")[0].getText())
        lowprice = Str2Float(price[2].select("td")[0].getText())
        endprice = Str2Float(price[3].select("td")[0].getText())

        # 寄りからの上昇率
        today_rising = (highprice - openprice) / yesterday_endprice * 100
        # S高マーク
        s_mark = price[1].select("td")[1].getText()
        # 始値から5%以上上昇orストップ高引け銘柄のみ抽出
        if (today_rising > 5) or (highprice == endprice and s_mark == "S"):
            # excelに保存する処理
            max_row = ws.max_row + 1
            # 書き込んだ行を出力
            print(max_row)
            ws.cell(row=max_row, column=1).value = datetime.datetime.now()
            ws.cell(row=max_row, column=2).value = str(meigaralists[i])
            ws.cell(row=max_row, column=3).value = int(yesterday_endprice)
            ws.cell(row=max_row, column=4).value = int(openprice)
            ws.cell(row=max_row, column=5).value = int(highprice)
            ws.cell(row=max_row, column=6).value = int(lowprice)
            ws.cell(row=max_row, column=7).value = int(endprice)
            ws.cell(row=max_row, column=8).value = str(descriptions[i])
            wb.save("kabutan.xlsx")
    # 例外処理
    except Exception as e:
        print(e.message)
        pass
